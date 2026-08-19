import csv
import io
import json
import math
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse, urlunparse

from openpyxl import load_workbook


ROOT = Path("/Users/svalenza/Documents/CEO Ask")
BOTIFY_ZIP = Path(
    "/Users/svalenza/Downloads/export_2072255_4294835_www.chewy.com_2026-08-18_15-07.zip"
)
BOTIFY_CSV = "export_2072255_4294835_www.chewy.com_2026-08-18_15-07.csv"
KEYWORD_XLSX = Path("/Users/svalenza/Downloads/Keywords_2026-08-18.xlsx")
OUT_DIR = ROOT / "outputs"
BOTIFY_AUDIENCE_KEYWORD_COLUMN = (
    "No. of Keywords for the URL To Achieve 90% Audience (by Country)"
)

PAGE_TYPES = [
    "Browse",
    "HVF",
    "HVSP",
    "Top Picks",
    "PDP",
    "Brand",
    "Editorial",
    "Deals",
]

SCHEMA_SIGNAL_COLUMNS = [
    "Product Exists",
    "Review Exists",
    "Rating Exists",
    "Offer Exists",
    "FAQ Exists",
    "Breadcrumb Exists",
]


def normalized_url(value):
    if value is None:
        return None
    raw = str(value).strip()
    if not raw or raw.upper() == "N/A":
        return None
    if "://" in raw:
        parsed = urlparse(raw)
    elif raw.startswith("/"):
        parsed = urlparse(f"https://www.chewy.com{raw}")
    else:
        parsed = urlparse(f"https://{raw}")
    path = parsed.path or "/"
    if len(path) > 1 and path.endswith("/"):
        path = path[:-1]
    return urlunparse(("https", parsed.netloc.lower(), path, "", "", ""))


def page_type_for(url):
    path = (urlparse(url or "").path.rstrip("/") or "/").lower()
    if path == "/deals" or path.startswith("/deals/"):
        return "Deals"
    if path.startswith("/best/"):
        return "Top Picks"
    if "/dp/" in path:
        return "PDP"
    if path.startswith("/f/"):
        return "HVF"
    if path.startswith("/sp/"):
        return "HVSP"
    if path == "/brands" or path.startswith("/brands/"):
        return "Brand"
    if path.startswith("/b/"):
        return "Browse"
    if (
        path == "/education"
        or path.startswith("/education/")
        or path.startswith("/bechewy/")
        or path.startswith("/petmd/")
    ):
        return "Editorial"
    return "Other"


def text(value):
    return (value or "").strip()


def parse_bool(value):
    lowered = text(value).lower()
    if lowered == "true":
        return True
    if lowered == "false":
        return False
    return None


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    raw = str(value).strip().replace(",", "")
    if not raw or raw.upper() == "N/A" or raw.lower() == "did not rank":
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def parse_int(value):
    parsed = parse_number(value)
    return int(parsed) if parsed is not None else None


def h1_values(value):
    raw = text(value)
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]
    except json.JSONDecodeError:
        pass
    return [raw]


def tokens(value):
    stopwords = {
        "www",
        "com",
        "the",
        "and",
        "for",
        "with",
        "from",
        "chewy",
        "free",
        "shipping",
        "best",
    }
    return [
        token
        for token in re.findall(r"[a-z0-9]+", (value or "").lower())
        if len(token) > 2 and token not in stopwords
    ]


SEMANTIC_ALIASES = {
    "cats": "cat",
    "dogs": "dog",
    "kittens": "kitten",
    "puppies": "puppy",
    "pets": "pet",
    "treats": "treat",
    "chews": "chew",
    "toys": "toy",
    "products": "product",
    "supplies": "supply",
    "supplements": "supplement",
    "veterinary": "prescription",
    "vet": "prescription",
    "rx": "prescription",
    "hills": "hill",
    "lickins": "lickin",
}

SEMANTIC_STOPWORDS = {
    "www",
    "com",
    "the",
    "and",
    "for",
    "with",
    "from",
    "free",
    "shipping",
    "chewy",
    "brand",
    "brands",
    "page",
    "pages",
    "shop",
    "buy",
    "online",
    "top",
    "best",
}

SPECIES_TOKENS = {
    "bird",
    "cat",
    "dog",
    "fish",
    "horse",
    "kitten",
    "pet",
    "puppy",
    "reptile",
    "rabbit",
}

GENERIC_BRAND_URL_TOKENS = {
    "adult",
    "allergy",
    "care",
    "cat",
    "chew",
    "coat",
    "deal",
    "diet",
    "dog",
    "dry",
    "food",
    "health",
    "kitten",
    "pet",
    "prescription",
    "product",
    "puppy",
    "skin",
    "supply",
    "treat",
    "toy",
    "wet",
}


def semantic_token(token):
    cleaned = token.lower().strip("'")
    if cleaned in SEMANTIC_ALIASES:
        return SEMANTIC_ALIASES[cleaned]
    if cleaned.endswith("ies") and len(cleaned) > 4:
        return f"{cleaned[:-3]}y"
    if cleaned.endswith("s") and len(cleaned) > 4 and not cleaned.endswith("ss"):
        return cleaned[:-1]
    return cleaned


def semantic_tokens(value):
    cleaned = str(value or "").replace("&", " and ")
    output = []
    for raw_token in re.findall(r"[a-z0-9]+", cleaned.lower()):
        if raw_token.isdigit() or re.fullmatch(r"[cfsv]\d+[a-z0-9]*", raw_token):
            continue
        token = semantic_token(raw_token)
        if len(token) < 2 or token in SEMANTIC_STOPWORDS:
            continue
        output.append(token)
    return output


def slug_for(url):
    parts = [part for part in urlparse(url or "").path.strip("/").split("/") if part]
    if not parts:
        return ""
    if "dp" in parts:
        index = parts.index("dp")
        return parts[index - 1] if index > 0 else ""
    return parts[-1]


def human_slug_for(url):
    slug = slug_for(url).lower()
    slug = re.sub(r"_(?:c\d+|f\d+[a-z0-9]*|s\d+).*$", "", slug)
    slug = re.sub(r"-\d+$", "", slug)
    return slug


def clean_url(url):
    parsed = urlparse(url or "")
    path = parsed.path
    return (
        bool(parsed.scheme and parsed.netloc)
        and not parsed.query
        and not parsed.fragment
        and path == path.lower()
        and "%20" not in path
        and "//" not in path
    )


def stable_url_id_points(row):
    url = row["_normalized_url"]
    path = urlparse(url).path.strip("/")
    parts = [part for part in path.split("/") if part]
    slug = slug_for(url).lower()
    page_type = row["_page_type"]
    if page_type == "PDP":
        return 10 if "/dp/" in urlparse(url).path and len(parts) >= 3 else 0
    if page_type in {"Browse", "Brand", "Deals"}:
        return 10 if re.search(r"-\d+$", slug) else 0
    if page_type == "HVF":
        return 10 if re.search(r"_c\d+_f\d*[a-z]?\d+$", slug) else 0
    if page_type == "Top Picks":
        return 10 if re.search(r"_s\d+$", slug) else 0
    if page_type in {"HVSP", "Editorial"}:
        return 10 if len(semantic_tokens(human_slug_for(url).replace("-", " "))) >= 1 else 0
    return 0


def target_text_for_url(row, main_keyword_by_url):
    main_keyword = main_keyword_by_url.get(row["_normalized_url"])
    if main_keyword:
        return main_keyword
    return f"{row.get('Title') or ''} {' '.join(h1_values(row.get('metadata-h1-contents')))}"


def url_intent_match_quality(row, main_keyword_by_url):
    target_text = target_text_for_url(row, main_keyword_by_url)
    target_tokens = set(semantic_tokens(target_text))
    slug_tokens = set(semantic_tokens(human_slug_for(row["_normalized_url"]).replace("-", " ")))
    if not target_tokens or not slug_tokens:
        return None
    coverage = len(slug_tokens & target_tokens) / len(target_tokens)
    if coverage >= 0.80:
        quality = "full"
    elif coverage >= 0.50:
        quality = "partial"
    elif coverage > 0:
        quality = "weak"
    else:
        quality = "fail"

    if row["_page_type"] in {"Browse", "Brand"}:
        target_species = target_tokens & SPECIES_TOKENS
        if target_species and not slug_tokens & target_species:
            return "weak" if quality in {"full", "partial"} else quality

    if row["_page_type"] == "Brand":
        target_brand_tokens = target_tokens - GENERIC_BRAND_URL_TOKENS
        if target_brand_tokens and not slug_tokens & target_brand_tokens:
            return "weak" if quality in {"full", "partial"} else quality

    return quality


def url_intent_points(row, main_keyword_by_url, max_points):
    quality = url_intent_match_quality(row, main_keyword_by_url)
    if quality == "full":
        return max_points
    if quality == "partial":
        return max_points * 0.50
    if quality == "weak":
        return max_points * 0.25
    if quality == "fail":
        return 0
    return None


def overlap_score(url, title, h1s):
    slug_tokens = set(tokens(slug_for(url).replace("-", " ")))
    visible_tokens = set(tokens(f"{title or ''} {' '.join(h1s)}"))
    if not slug_tokens:
        return 0
    ratio = len(slug_tokens & visible_tokens) / len(slug_tokens)
    if ratio >= 0.60:
        return 20
    if ratio >= 0.35:
        return 12
    if ratio > 0:
        return 6
    return 0


def metric_quality(score_text, aggregate, metric):
    if metric == "LCP":
        if aggregate is None:
            return metric_quality_from_label(score_text)
        if aggregate < 2500:
            return "good"
        return "needs" if aggregate <= 4000 else "poor"
    if metric == "CLS":
        if aggregate is None:
            return metric_quality_from_label(score_text)
        if aggregate < 0.1:
            return "good"
        return "needs" if aggregate <= 0.25 else "poor"
    if metric == "INP":
        if aggregate is None:
            return metric_quality_from_label(score_text)
        if aggregate < 200:
            return "good"
        return "needs" if aggregate <= 500 else "poor"
    if metric == "FID":
        if aggregate is None:
            return metric_quality_from_label(score_text)
        if aggregate <= 100:
            return "good"
        return "needs" if aggregate <= 300 else "poor"
    return metric_quality_from_label(score_text)


def metric_quality_from_label(score_text):
    score = text(score_text).lower()
    if score.startswith("good"):
        return "good"
    if score.startswith("need"):
        return "needs"
    if score.startswith("poor"):
        return "poor"
    return None


def is_bad_crawl(row):
    has_extracted_text = bool(
        text(row.get("Title"))
        or text(row.get("Meta Description"))
        or h1_values(row.get("metadata-h1-contents"))
    )
    has_schema_signal = any(
        parse_bool(row.get(column)) is True for column in SCHEMA_SIGNAL_COLUMNS
    )
    has_no_data_marker = any(
        text(value).lower() == "no data" for value in row.values()
    )
    has_no_onpage_data = (
        not has_extracted_text
        and not has_schema_signal
        and (parse_int(row.get("No. of Unique Outlinks to Internal Pages")) or 0) == 0
    )
    has_index_signal = (
        parse_bool(row.get("Canonical Points to Self")) is True
        or parse_bool(row.get("In Sitemaps")) is True
    )
    return has_no_data_marker or has_no_onpage_data or (
        not has_extracted_text and not has_schema_signal and not has_index_signal
    )


def read_botify():
    rows = []
    title_counts = Counter()
    meta_counts = Counter()
    h1_counts = Counter()
    with zipfile.ZipFile(BOTIFY_ZIP) as archive:
        with archive.open(BOTIFY_CSV) as handle:
            stream = io.TextIOWrapper(handle, encoding="utf-8-sig", newline="")
            stream.readline()
            reader = csv.DictReader(stream)
            for row in reader:
                url = normalized_url(row.get("Full URL"))
                if not url:
                    continue
                row["_normalized_url"] = url
                row["_page_type"] = page_type_for(url)
                rows.append(row)
                title = text(row.get("Title")).lower()
                meta = text(row.get("Meta Description")).lower()
                h1_key = " | ".join(h1_values(row.get("metadata-h1-contents"))).lower()
                if title:
                    title_counts[title] += 1
                if meta:
                    meta_counts[meta] += 1
                if h1_key:
                    h1_counts[h1_key] += 1
    return rows, title_counts, meta_counts, h1_counts


def keyword_match_quality(container, keyword):
    if not keyword:
        return None
    container_text = text(container).lower()
    keyword_text = text(keyword).lower()
    if not container_text or not keyword_text:
        return "fail"
    if keyword_text in container_text:
        return "full"
    keyword_tokens = set(tokens(keyword_text))
    container_tokens = set(tokens(container_text))
    if not keyword_tokens:
        return None
    coverage = len(keyword_tokens & container_tokens) / len(keyword_tokens)
    if coverage >= 0.80:
        return "full"
    if coverage >= 0.50:
        return "partial"
    return "fail"


def keyword_match_points(container, keyword, max_points):
    quality = keyword_match_quality(container, keyword)
    if quality == "full":
        return max_points
    if quality == "partial":
        return max_points * 0.50
    if quality == "fail":
        return 0
    return None


def semantic_match_quality(container, target):
    if not target:
        return None
    container_text = text(container)
    target_text = text(target)
    if not container_text or not target_text:
        return "fail"
    container_tokens = set(semantic_tokens(container_text))
    target_tokens = set(semantic_tokens(target_text))
    if not target_tokens:
        return None
    coverage = len(container_tokens & target_tokens) / len(target_tokens)
    if coverage >= 0.80:
        return "full"
    if coverage >= 0.50:
        return "partial"
    if coverage > 0:
        return "weak"
    return "fail"


def quality_points(quality, max_points):
    if quality == "full":
        return max_points
    if quality == "partial":
        return max_points * 0.65
    if quality == "weak":
        return max_points * 0.30
    if quality == "fail":
        return 0
    return None


def meta_base_score(row, title_counts, meta_counts):
    title = text(row.get("Title"))
    meta = text(row.get("Meta Description"))
    title_len = parse_int(row.get("Title Length")) or len(title)
    meta_len = parse_int(row.get("Meta Description Length")) or len(meta)
    score = 0
    if title:
        score += 25
        if 30 <= title_len <= 65:
            score += 15
        elif 20 <= title_len <= 75:
            score += 8
        duplicates = title_counts[title.lower()]
        score += 10 if duplicates == 1 else 5 if duplicates <= 10 else 0
    if meta:
        score += 25
        if 70 <= meta_len <= 170:
            score += 15
        elif 50 <= meta_len <= 200:
            score += 8
        duplicates = meta_counts[meta.lower()]
        score += 10 if duplicates == 1 else 5 if duplicates <= 10 else 0
    return score


def meta_score(row, title_counts, meta_counts, main_keyword_by_url):
    base_score = meta_base_score(row, title_counts, meta_counts)
    main_keyword = main_keyword_by_url.get(row["_normalized_url"])
    if not main_keyword:
        return base_score
    title_points = keyword_match_points(row.get("Title"), main_keyword, 15)
    meta_points = keyword_match_points(row.get("Meta Description"), main_keyword, 10)
    return base_score * 0.75 + (title_points or 0) + (meta_points or 0)


def legacy_schema_proxy(row):
    checks = [
        ("Breadcrumb Exists", 25),
        ("Product Exists", 20),
        ("Offer Exists", 15),
        ("Rating Exists", 15),
        ("Review Exists", 10),
        ("FAQ Exists", 10),
    ]
    score = sum(weight for column, weight in checks if parse_bool(row.get(column)) is True)
    if (parse_int(row.get("No. of Reviews")) or 0) > 0 or (
        parse_int(row.get("No. of Questions")) or 0
    ) > 0:
        score += 5
    return min(100, score)


def schema_score(row):
    page_type = row["_page_type"]
    breadcrumb = parse_bool(row.get("Breadcrumb Exists")) is True
    faq = parse_bool(row.get("FAQ Exists")) is True
    if page_type == "PDP":
        return (
            (35 if parse_bool(row.get("Product Exists")) is True else 0)
            + (30 if parse_bool(row.get("Offer Exists")) is True else 0)
            + (25 if breadcrumb else 0)
            + (5 if parse_bool(row.get("Rating Exists")) is True else 0)
            + (5 if faq else 0)
        )
    if page_type in {"Browse", "HVF"}:
        return (80 if breadcrumb else 0) + (20 if faq else 0)
    if page_type == "Editorial":
        return 90
    if page_type in {"HVSP", "Top Picks", "Brand", "Deals"}:
        return 100 if breadcrumb else 0
    return legacy_schema_proxy(row)


def url_score(row, main_keyword_by_url):
    url = row["_normalized_url"]
    score = 0
    if row["_page_type"] != "Other":
        score += 25
    if clean_url(url):
        score += 15
    score += stable_url_id_points(row)
    slug_tokens = semantic_tokens(human_slug_for(url).replace("-", " "))
    if len(slug_tokens) >= 2:
        score += 20
    elif len(slug_tokens) == 1:
        score += 12
    intent_points = url_intent_points(row, main_keyword_by_url, 25)
    if intent_points is not None:
        score += intent_points
    elif overlap_score(url, row.get("Title"), h1_values(row.get("metadata-h1-contents"))) >= 12:
        score += 12
    if len(urlparse(url).path) <= 115 and len(
        [part for part in urlparse(url).path.split("/") if part]
    ) <= 4:
        score += 5
    return min(100, score)


def h1_text_for(row):
    return " ".join(h1_values(row.get("metadata-h1-contents")))


def title_h1_match_quality(row):
    return semantic_match_quality(row.get("Title"), h1_text_for(row))


def h1_target_match_quality(row, main_keyword_by_url):
    main_keyword = main_keyword_by_url.get(row["_normalized_url"])
    if main_keyword:
        return semantic_match_quality(h1_text_for(row), main_keyword)
    return title_h1_match_quality(row)


def h1_support_points(row):
    page_type = row["_page_type"]
    audience_signal = (parse_int(row.get(BOTIFY_AUDIENCE_KEYWORD_COLUMN)) or 0) > 0
    product_signal = parse_bool(row.get("Product Exists")) is True
    offer_signal = parse_bool(row.get("Offer Exists")) is True
    review_signal = parse_bool(row.get("Review Exists")) is True
    rating_signal = parse_bool(row.get("Rating Exists")) is True
    faq_signal = parse_bool(row.get("FAQ Exists")) is True
    reviews_or_questions = (
        (parse_int(row.get("No. of Reviews")) or 0) > 0
        or (parse_int(row.get("No. of Questions")) or 0) > 0
    )

    if page_type == "PDP":
        return (
            (8 if product_signal else 0)
            + (4 if reviews_or_questions or review_signal or rating_signal else 0)
            + (3 if audience_signal else 0)
        )
    if page_type == "Editorial":
        return 10 if audience_signal else 0
    if page_type == "Deals":
        return (
            (8 if offer_signal or product_signal or review_signal else 0)
            + (7 if audience_signal else 0)
        )
    if page_type in {"Browse", "HVF", "HVSP", "Top Picks", "Brand"}:
        return (
            (8 if product_signal or review_signal or rating_signal or faq_signal else 0)
            + (7 if audience_signal else 0)
        )
    return 7 if audience_signal else 0


def h1_score(row, h1_counts, main_keyword_by_url):
    h1s = h1_values(row.get("metadata-h1-contents"))
    score = 0
    if h1s:
        score += 25
    if len(h1s) == 1:
        score += 10
    elif len(h1s) > 1:
        score += 5
    h1_token_count = len(semantic_tokens(" ".join(h1s)))
    if h1_token_count >= 2:
        score += 10
    elif h1_token_count == 1:
        score += 5

    target_quality = h1_target_match_quality(row, main_keyword_by_url)
    target_points = quality_points(target_quality, 25)
    if target_points is not None:
        score += target_points

    consistency_points = quality_points(title_h1_match_quality(row), 10)
    if consistency_points is not None:
        score += consistency_points

    score += h1_support_points(row)

    h1_key = " | ".join(h1s).lower()
    if h1_key:
        duplicates = h1_counts[h1_key]
        score += 5 if duplicates <= 5 else 3 if duplicates <= 20 else 0
    return min(100, score)


def noindex_enabled(row):
    return parse_bool(row.get("X-Robots-Tags: Noindex")) is True


def nofollow_enabled(row):
    return parse_bool(row.get("X-Robots-Tags: Nofollow")) is True


def self_canonical_enabled(row):
    return parse_bool(row.get("Canonical Points to Self")) is True


def sitemap_inclusion_enabled(row):
    return parse_bool(row.get("In Sitemaps")) is True


def sitemap_compliant(row):
    if noindex_enabled(row):
        return not sitemap_inclusion_enabled(row)
    return sitemap_inclusion_enabled(row)


def canonical_sitemap_compliant(row):
    if noindex_enabled(row):
        return sitemap_compliant(row)
    return self_canonical_enabled(row) and sitemap_compliant(row)


def index_score(row):
    if noindex_enabled(row):
        score = 50
        if not nofollow_enabled(row):
            score += 10
        if self_canonical_enabled(row):
            score += 20
        if sitemap_compliant(row):
            score += 20
        return min(100, score)

    score = 0
    score += 20
    if not nofollow_enabled(row):
        score += 10
    if self_canonical_enabled(row):
        score += 40
    if sitemap_compliant(row):
        score += 30
    if nofollow_enabled(row):
        score = min(score, 85)
    if not self_canonical_enabled(row):
        score = min(score, 60)
    if not sitemap_compliant(row):
        score = min(score, 70)
    return min(100, score)


INLINK_TARGETS = {
    "Browse": (5, 25, 75),
    "HVF": (2, 10, 30),
    "HVSP": (5, 20, 50),
    "Top Picks": (5, 20, 50),
    "PDP": (1, 5, 15),
    "Brand": (2, 8, 25),
    "Editorial": (1, 5, 15),
    "Deals": (5, 25, 50),
    "Other": (1, 3, 10),
}
OUTLINK_TARGETS = {
    "Browse": (10, 50, 150),
    "HVF": (10, 50, 150),
    "HVSP": (10, 50, 150),
    "Top Picks": (10, 50, 150),
    "PDP": (5, 20, 100),
    "Brand": (5, 20, 100),
    "Editorial": (5, 20, 100),
    "Deals": (10, 50, 150),
    "Other": (1, 10, 30),
}
DEPTH_TARGETS = {
    "Browse": (3, 5),
    "HVF": (3, 5),
    "HVSP": (3, 5),
    "Top Picks": (3, 5),
    "PDP": (5, 8),
    "Brand": (3, 5),
    "Editorial": (3, 5),
    "Deals": (3, 5),
    "Other": (3, 5),
}


def tier_points(value, low, medium, high, max_points):
    if value is None:
        return 0
    if value >= high:
        return max_points
    if value >= medium:
        return max_points * 0.65
    if value >= low:
        return max_points * 0.30
    return 0


def depth_points(row, max_points):
    depth = parse_int(row.get("Depth"))
    if depth is None:
        return 0
    full, partial = DEPTH_TARGETS.get(row["_page_type"], DEPTH_TARGETS["Other"])
    if depth <= full:
        return max_points
    if depth <= partial:
        return max_points * 0.60
    if depth <= 8:
        return max_points * 0.20
    return 0


def depth_ok(row):
    depth = parse_int(row.get("Depth"))
    if depth is None:
        return False
    _, partial = DEPTH_TARGETS.get(row["_page_type"], DEPTH_TARGETS["Other"])
    return depth <= partial


def inlink_target_met(row):
    value = parse_int(row.get("No. of Unique Inlinks")) or 0
    _, medium, _ = INLINK_TARGETS.get(row["_page_type"], INLINK_TARGETS["Other"])
    return value >= medium


def outlink_target_met(row):
    value = parse_int(row.get("No. of Unique Outlinks to Internal Pages")) or 0
    _, medium, _ = OUTLINK_TARGETS.get(row["_page_type"], OUTLINK_TARGETS["Other"])
    return value >= medium


def links_score(row):
    page_type = row["_page_type"]
    score = tier_points(
        parse_int(row.get("No. of Unique Inlinks")),
        *INLINK_TARGETS.get(page_type, INLINK_TARGETS["Other"]),
        45,
    )
    score += depth_points(row, 25)
    score += tier_points(
        parse_int(row.get("No. of Unique Outlinks to Internal Pages")),
        *OUTLINK_TARGETS.get(page_type, OUTLINK_TARGETS["Other"]),
        20,
    )
    if parse_bool(row.get("Breadcrumb Exists")) is True:
        score += 10
    if (parse_int(row.get("No. of Unique Inlinks")) or 0) == 0:
        score = min(score, 60)
    if (parse_int(row.get("No. of Unique Outlinks to Internal Pages")) or 0) == 0:
        score = min(score, 75)
    return min(100, score)


def metric_points(quality):
    if quality == "good":
        return 100
    if quality == "needs":
        return 70
    if quality == "poor":
        return 0
    return None


def row_metric_quality(row, metric):
    return metric_quality(
        row.get(f"{metric}: Score"), parse_number(row.get(f"{metric}: Agg")), metric
    )


def row_has_metric(row, metric):
    return row_metric_quality(row, metric) is not None


def speed_available_metrics(row):
    return [metric for metric in ("LCP", "CLS", "INP") if row_has_metric(row, metric)]


def speed_measured_health(row):
    values = []
    for metric in speed_available_metrics(row):
        points = metric_points(row_metric_quality(row, metric))
        values.append(points)
    return sum(values) / len(values) if values else None


def rank_bucket_points(rank):
    if rank is None:
        return 0
    if rank <= 3:
        return 100
    if rank <= 10:
        return 85
    if rank <= 20:
        return 70
    if rank <= 50:
        return 45
    if rank <= 100:
        return 20
    return 0


def main_keyword_candidate_key(record):
    rank = parse_number(record.get("Rank Current"))
    msv = parse_number(record.get("MSV")) or 0
    rank_sort = 9999 if rank is None else rank
    return (
        1 if text(record.get("Prime")).upper() == "Y" else 0,
        1 if text(record.get("Is Highest Ranking")).upper() == "Y" else 0,
        1 if text(record.get("Highest Ranking")).upper() == "Y" else 0,
        msv,
        -rank_sort,
    )


def read_keywords():
    headers = [
        "Keyword",
        "MSV",
        "Tracking",
        "Rank Previous",
        "Rank Current",
        "Rank Change",
        "Result Type Previous",
        "Result Type Current",
        "Ranking URL",
        "Title",
        "Highest Ranking",
        "Preferred URL",
        "Is Highest Ranking",
        "Prime",
        "Keyword Groups",
        "Website",
        "Search Engine",
        "Device",
        "Location",
    ]
    workbook = load_workbook(KEYWORD_XLSX, read_only=True, data_only=True)
    sheet = workbook["Expanded Keyword Performance"]
    main_keyword_candidates = {}
    keyword_by_url = defaultdict(
        lambda: {
            "keywords": set(),
            "rows": 0,
            "current_ranked": 0,
            "lost": 0,
            "top3": 0,
            "top10": 0,
            "top20": 0,
            "top50": 0,
            "top100": 0,
            "score_sum": 0.0,
            "rank_sum": 0.0,
            "rank_n": 0,
            "groups": Counter(),
        }
    )
    for row in sheet.iter_rows(min_row=8, values_only=True):
        record = dict(zip(headers, row[: len(headers)]))
        keyword = text(record.get("Keyword"))
        url = normalized_url(record.get("Ranking URL"))
        if not keyword or not url:
            continue
        candidate_key = main_keyword_candidate_key(record)
        if url not in main_keyword_candidates or candidate_key > main_keyword_candidates[url][0]:
            main_keyword_candidates[url] = (candidate_key, keyword)
        rank = parse_number(record.get("Rank Current"))
        data = keyword_by_url[url]
        data["keywords"].add(keyword.lower())
        data["rows"] += 1
        data["score_sum"] += rank_bucket_points(rank)
        if rank is None:
            data["lost"] += 1
        else:
            data["current_ranked"] += 1
            data["rank_sum"] += rank
            data["rank_n"] += 1
            if rank <= 3:
                data["top3"] += 1
            if rank <= 10:
                data["top10"] += 1
            if rank <= 20:
                data["top20"] += 1
            if rank <= 50:
                data["top50"] += 1
            if rank <= 100:
                data["top100"] += 1
        for group in str(record.get("Keyword Groups") or "").split(","):
            cleaned = group.strip()
            if cleaned:
                data["groups"][cleaned] += 1
    main_keyword_by_url = {
        url: keyword for url, (_, keyword) in main_keyword_candidates.items()
    }
    return keyword_by_url, main_keyword_by_url


def keyword_score(data):
    rows = data["rows"]
    if not rows:
        return None
    quality = data["score_sum"] / rows
    top10 = data["top10"] / rows * 100
    active = data["current_ranked"] / rows * 100
    return min(100, 0.75 * quality + 0.15 * top10 + 0.10 * active)


def keyword_opportunity_score(data):
    rows = data["rows"]
    if not rows:
        return None
    top10_gap = (rows - data["top10"]) / rows * 100
    quick_win = max(0, data["top20"] - data["top10"]) / rows * 100
    striking_distance = max(0, data["top50"] - data["top20"]) / rows * 100
    lost_or_unranked = data["lost"] / rows * 100
    return min(
        100,
        0.45 * top10_gap
        + 0.30 * quick_win
        + 0.15 * striking_distance
        + 0.10 * lost_or_unranked,
    )


def botify_audience_potential(row):
    return parse_number(row.get(BOTIFY_AUDIENCE_KEYWORD_COLUMN))


def numeric_distribution(values):
    sorted_values = sorted(value for value in values if value is not None)
    if not sorted_values:
        return {"count": 0, "total": 0, "avg": None, "median": None, "p90": None}
    middle = len(sorted_values) // 2
    median = (
        sorted_values[middle]
        if len(sorted_values) % 2
        else (sorted_values[middle - 1] + sorted_values[middle]) / 2
    )
    p90 = sorted_values[int((len(sorted_values) - 1) * 0.9)]
    total = sum(sorted_values)
    return {
        "count": len(sorted_values),
        "total": total,
        "avg": total / len(sorted_values),
        "median": median,
        "p90": p90,
    }


def log_index(value, max_value):
    if not value or not max_value:
        return 0
    return math.log1p(value) / math.log1p(max_value) * 100


def ranking_potential_score(summary, audience_summary, maxes):
    rows = summary["rows"]
    if not rows and not audience_summary["count"]:
        return None
    top10_gap = (rows - summary["top10"]) / rows * 100 if rows else 0
    quick_win = (
        max(0, summary["top20"] - summary["top10"]) / rows * 100 if rows else 0
    )
    striking_distance = (
        max(0, summary["top50"] - summary["top20"]) / rows * 100 if rows else 0
    )
    quick_win_index = quick_win / maxes["quick_win"] * 100 if maxes["quick_win"] else 0
    striking_index = (
        striking_distance / maxes["striking_distance"] * 100
        if maxes["striking_distance"]
        else 0
    )
    audience_intensity_index = (
        audience_summary["avg"] / maxes["avg_audience"] * 100
        if audience_summary["avg"] is not None and maxes["avg_audience"]
        else 0
    )
    audience_total_index = log_index(
        audience_summary["total"], maxes["total_audience"]
    )
    return min(
        100,
        0.30 * top10_gap
        + 0.25 * quick_win_index
        + 0.15 * striking_index
        + 0.15 * audience_intensity_index
        + 0.15 * audience_total_index,
    )


def average(rows, score_fn):
    values = [score_fn(row) for row in rows]
    values = [value for value in values if value is not None]
    return sum(values) / len(values) if values else None


def median_numeric(rows, column):
    values = sorted(
        value
        for value in (parse_int(row.get(column)) for row in rows)
        if value is not None
    )
    if not values:
        return None
    middle = len(values) // 2
    if len(values) % 2:
        return values[middle]
    return (values[middle - 1] + values[middle]) / 2


def pct_rows(rows, predicate):
    if not rows:
        return 0
    return round(sum(1 for row in rows if predicate(row)) / len(rows) * 100, 1)


def metric_measured_pct(rows, metric):
    return pct_rows(rows, lambda row: row_has_metric(row, metric))


def metric_good_pct(rows, metric):
    measured = [row for row in rows if row_has_metric(row, metric)]
    if not measured:
        return ""
    return round(
        sum(1 for row in measured if row_metric_quality(row, metric) == "good")
        / len(measured)
        * 100,
        1,
    )


ISSUE_FIELDS = [
    "page_type",
    "category",
    "issue",
    "affected_urls",
    "eligible_urls",
    "affected_pct",
    "severity",
    "evidence",
    "recommendation",
    "detail_file",
]

ISSUE_DETAIL_FIELDS = [
    "page_type",
    "category",
    "issue",
    "normalized_url",
    "matched_botify",
    "recommendation",
    "evidence_value",
    "title",
    "title_length",
    "meta_description",
    "meta_description_length",
    "main_keyword",
    "h1",
    "h1_count",
    "url_slug",
    "canonical_points_to_self",
    "in_sitemaps",
    "noindex",
    "nofollow",
    "breadcrumb_exists",
    "product_exists",
    "offer_exists",
    "faq_exists",
    "unique_inlinks",
    "unique_outlinks",
    "depth",
    "lcp_status",
    "lcp_value",
    "cls_status",
    "cls_value",
    "inp_status",
    "inp_value",
    "ranking_keyword_rows",
    "distinct_keywords",
    "avg_current_rank",
    "top10_keywords",
    "quick_win_11_20_keywords",
    "striking_distance_21_50_keywords",
    "lost_or_unranked_keywords",
    "botify_audience_keywords_to_90pct",
]

CATEGORY_ORDER = {
    "Meta": 1,
    "Schema": 2,
    "URL": 3,
    "H1": 4,
    "Index": 5,
    "Links": 6,
    "Speed": 7,
    "Ranking Keywords": 8,
    "Ranking Potential": 9,
}


def issue_severity(affected_count, eligible_count, severity=None):
    if severity:
        return severity
    if not eligible_count:
        return "low"
    pct = affected_count / eligible_count * 100
    if pct >= 50:
        return "high"
    if pct >= 20:
        return "medium"
    return "low"


def add_issue(
    issue_rows,
    page_type,
    category,
    issue,
    affected_count,
    eligible_count,
    evidence,
    recommendation,
    severity=None,
):
    if affected_count <= 0 or eligible_count <= 0:
        return
    issue_rows.append(
        {
            "page_type": page_type,
            "category": category,
            "issue": issue,
            "affected_urls": affected_count,
            "eligible_urls": eligible_count,
            "affected_pct": round(affected_count / eligible_count * 100, 1),
            "severity": issue_severity(affected_count, eligible_count, severity),
            "evidence": evidence,
            "recommendation": recommendation,
            "detail_file": "",
        }
    )


def count_rows(rows, predicate):
    return sum(1 for row in rows if predicate(row))


def title_length(row):
    title = text(row.get("Title"))
    return parse_int(row.get("Title Length")) or len(title)


def meta_description_length(row):
    meta_description = text(row.get("Meta Description"))
    return parse_int(row.get("Meta Description Length")) or len(meta_description)


def path_too_long_or_deep(row):
    url = row["_normalized_url"]
    return len(urlparse(url).path) > 115 or len(
        [part for part in urlparse(url).path.split("/") if part]
    ) > 4


def keyword_urls_for_page_type(keyword_by_url, page_type):
    return {
        url: data
        for url, data in keyword_by_url.items()
        if page_type_for(url) == page_type
    }


def issue_slug(*parts):
    raw = "-".join(text(part).lower() for part in parts)
    slug = re.sub(r"[^a-z0-9]+", "-", raw).strip("-")
    return slug[:140] or "issue"


def issue_detail_filename(page_type, category, issue):
    return f"{issue_slug(page_type, category, issue)}.csv"


def keyword_quick_win_count(data):
    return max(0, data.get("top20", 0) - data.get("top10", 0))


def keyword_striking_distance_count(data):
    return max(0, data.get("top50", 0) - data.get("top20", 0))


def keyword_avg_rank(data):
    return data["rank_sum"] / data["rank_n"] if data.get("rank_n") else ""


def keyword_detail_values(data):
    if not data:
        return {
            "ranking_keyword_rows": "",
            "distinct_keywords": "",
            "avg_current_rank": "",
            "top10_keywords": "",
            "quick_win_11_20_keywords": "",
            "striking_distance_21_50_keywords": "",
            "lost_or_unranked_keywords": "",
        }
    return {
        "ranking_keyword_rows": data["rows"],
        "distinct_keywords": len(data["keywords"]),
        "avg_current_rank": (
            "" if keyword_avg_rank(data) == "" else round(keyword_avg_rank(data), 2)
        ),
        "top10_keywords": data["top10"],
        "quick_win_11_20_keywords": keyword_quick_win_count(data),
        "striking_distance_21_50_keywords": keyword_striking_distance_count(data),
        "lost_or_unranked_keywords": data["lost"],
    }


def metric_value(row, metric):
    parsed = parse_number(row.get(f"{metric}: Agg"))
    return "" if parsed is None else parsed


def issue_evidence_value(row, issue, main_keyword_by_url):
    if issue == "Page title length is not ideal":
        return title_length(row)
    if issue == "Meta description length is not ideal":
        return meta_description_length(row)
    if issue == "Duplicate page title":
        return text(row.get("Title"))
    if issue == "Duplicate meta description":
        return text(row.get("Meta Description"))
    if issue in {
        "Title does not align to main ranking keyword",
        "Meta description does not align to main ranking keyword",
        "URL slug does not align to main keyword",
        "H1 does not align to main keyword or topic",
    }:
        return main_keyword_by_url.get(row["_normalized_url"], "")
    if issue in {"Missing H1", "Multiple H1s present", "H1 is too generic", "Duplicate H1 pattern"}:
        return h1_text_for(row)
    if issue in {"No unique inlinks", "Below inlink target"}:
        return parse_int(row.get("No. of Unique Inlinks")) or 0
    if issue in {"No internal outlinks", "Below outlink target"}:
        return parse_int(row.get("No. of Unique Outlinks to Internal Pages")) or 0
    if issue == "Crawl depth too deep":
        return parse_int(row.get("Depth")) or ""
    if issue.startswith(("LCP", "CLS", "INP")):
        metric = issue.split(" ", 1)[0]
        return metric_value(row, metric)
    if issue == "Botify audience-keyword potential present":
        return botify_audience_potential(row) or ""
    return ""


def botify_detail_row(
    row,
    category,
    issue,
    recommendation,
    keyword_by_url,
    main_keyword_by_url,
):
    url = row["_normalized_url"]
    keyword_data = keyword_by_url.get(url)
    details = {
        "page_type": row["_page_type"],
        "category": category,
        "issue": issue,
        "normalized_url": url,
        "matched_botify": True,
        "recommendation": recommendation,
        "evidence_value": issue_evidence_value(row, issue, main_keyword_by_url),
        "title": text(row.get("Title")),
        "title_length": title_length(row) if text(row.get("Title")) else "",
        "meta_description": text(row.get("Meta Description")),
        "meta_description_length": (
            meta_description_length(row) if text(row.get("Meta Description")) else ""
        ),
        "main_keyword": main_keyword_by_url.get(url, ""),
        "h1": h1_text_for(row),
        "h1_count": len(h1_values(row.get("metadata-h1-contents"))),
        "url_slug": human_slug_for(url),
        "canonical_points_to_self": text(row.get("Canonical Points to Self")),
        "in_sitemaps": text(row.get("In Sitemaps")),
        "noindex": text(row.get("X-Robots-Tags: Noindex")),
        "nofollow": text(row.get("X-Robots-Tags: Nofollow")),
        "breadcrumb_exists": text(row.get("Breadcrumb Exists")),
        "product_exists": text(row.get("Product Exists")),
        "offer_exists": text(row.get("Offer Exists")),
        "faq_exists": text(row.get("FAQ Exists")),
        "unique_inlinks": parse_int(row.get("No. of Unique Inlinks")) or 0,
        "unique_outlinks": parse_int(row.get("No. of Unique Outlinks to Internal Pages"))
        or 0,
        "depth": parse_int(row.get("Depth")) or "",
        "lcp_status": row_metric_quality(row, "LCP") or "",
        "lcp_value": metric_value(row, "LCP"),
        "cls_status": row_metric_quality(row, "CLS") or "",
        "cls_value": metric_value(row, "CLS"),
        "inp_status": row_metric_quality(row, "INP") or "",
        "inp_value": metric_value(row, "INP"),
        "botify_audience_keywords_to_90pct": botify_audience_potential(row) or "",
    }
    details.update(keyword_detail_values(keyword_data))
    return details


def keyword_issue_evidence_value(issue, data, matched_botify):
    if issue == "Ranking URL is not in Botify crawl":
        return "not matched" if not matched_botify else "matched"
    if issue == "No top-10 ranking keywords":
        return data["top10"]
    if issue == "Has quick-win keywords in positions 11-20":
        return keyword_quick_win_count(data)
    if issue == "Has striking-distance keywords in positions 21-50":
        return keyword_striking_distance_count(data)
    if issue == "Has lost or unranked keywords":
        return data["lost"]
    return ""


def keyword_detail_row(
    url,
    data,
    category,
    issue,
    recommendation,
    botify_by_url,
    main_keyword_by_url,
):
    botify_row = botify_by_url.get(url)
    if botify_row:
        detail = botify_detail_row(
            botify_row,
            category,
            issue,
            recommendation,
            {url: data},
            main_keyword_by_url,
        )
    else:
        detail = {field: "" for field in ISSUE_DETAIL_FIELDS}
        detail.update(
            {
                "page_type": page_type_for(url),
                "category": category,
                "issue": issue,
                "normalized_url": url,
                "matched_botify": False,
                "recommendation": recommendation,
                "main_keyword": main_keyword_by_url.get(url, ""),
            }
        )
        detail.update(keyword_detail_values(data))
    detail["evidence_value"] = keyword_issue_evidence_value(issue, data, bool(botify_row))
    return detail


def issue_detail_rows(
    summary_row,
    botify_by_type,
    keyword_by_url,
    botify_urls,
    title_counts,
    meta_counts,
    h1_counts,
    main_keyword_by_url,
):
    page_type = summary_row["page_type"]
    category = summary_row["category"]
    issue = summary_row["issue"]
    recommendation = summary_row["recommendation"]
    rows = [
        row for row in botify_by_type.get(page_type, []) if not is_bad_crawl(row)
    ]
    targeted_rows = [
        row for row in rows if row["_normalized_url"] in main_keyword_by_url
    ]
    botify_by_url = {
        row["_normalized_url"]: row
        for page_rows in botify_by_type.values()
        for row in page_rows
        if not is_bad_crawl(row)
    }

    schema_issue_columns = {
        "Product schema missing": "Product Exists",
        "Offer schema missing": "Offer Exists",
        "Breadcrumb schema missing": "Breadcrumb Exists",
        "Rating schema missing": "Rating Exists",
        "FAQ schema missing": "FAQ Exists",
        "Breadcrumb/navigation signal missing": "Breadcrumb Exists",
    }

    if issue == "Missing page title":
        matched_rows = [row for row in rows if not text(row.get("Title"))]
    elif issue == "Page title length is not ideal":
        matched_rows = [
            row
            for row in rows
            if text(row.get("Title")) and not (30 <= title_length(row) <= 65)
        ]
    elif issue == "Duplicate page title":
        matched_rows = [
            row
            for row in rows
            if text(row.get("Title"))
            and title_counts[text(row.get("Title")).lower()] > 1
        ]
    elif issue == "Missing meta description":
        matched_rows = [row for row in rows if not text(row.get("Meta Description"))]
    elif issue == "Meta description length is not ideal":
        matched_rows = [
            row
            for row in rows
            if text(row.get("Meta Description"))
            and not (70 <= meta_description_length(row) <= 170)
        ]
    elif issue == "Duplicate meta description":
        matched_rows = [
            row
            for row in rows
            if text(row.get("Meta Description"))
            and meta_counts[text(row.get("Meta Description")).lower()] > 1
        ]
    elif issue == "Title does not align to main ranking keyword":
        matched_rows = [
            row
            for row in targeted_rows
            if keyword_match_quality(
                row.get("Title"), main_keyword_by_url.get(row["_normalized_url"])
            )
            not in {"full", "partial"}
        ]
    elif issue == "Meta description does not align to main ranking keyword":
        matched_rows = [
            row
            for row in targeted_rows
            if keyword_match_quality(
                row.get("Meta Description"),
                main_keyword_by_url.get(row["_normalized_url"]),
            )
            not in {"full", "partial"}
        ]
    elif issue in schema_issue_columns:
        column = schema_issue_columns[issue]
        matched_rows = [row for row in rows if parse_bool(row.get(column)) is not True]
    elif issue == "Editorial schema implementation needs validation":
        matched_rows = rows
    elif issue == "URL does not match the expected stable ID pattern":
        matched_rows = [row for row in rows if stable_url_id_points(row) < 10]
    elif issue == "URL slug is too generic":
        matched_rows = [
            row
            for row in rows
            if len(
                semantic_tokens(human_slug_for(row["_normalized_url"]).replace("-", " "))
            )
            < 2
        ]
    elif issue == "URL slug does not align to main keyword":
        matched_rows = [
            row
            for row in targeted_rows
            if url_intent_match_quality(row, main_keyword_by_url)
            not in {"full", "partial"}
        ]
    elif issue == "URL path is too long or deep":
        matched_rows = [row for row in rows if path_too_long_or_deep(row)]
    elif issue == "Missing H1":
        matched_rows = [row for row in rows if not h1_values(row.get("metadata-h1-contents"))]
    elif issue == "Multiple H1s present":
        matched_rows = [
            row for row in rows if len(h1_values(row.get("metadata-h1-contents"))) > 1
        ]
    elif issue == "H1 is too generic":
        matched_rows = [
            row
            for row in rows
            if h1_values(row.get("metadata-h1-contents"))
            and len(semantic_tokens(h1_text_for(row))) < 2
        ]
    elif issue == "H1 does not align to main keyword or topic":
        matched_rows = [
            row
            for row in rows
            if h1_target_match_quality(row, main_keyword_by_url)
            not in {"full", "partial"}
        ]
    elif issue == "Duplicate H1 pattern":
        matched_rows = [
            row
            for row in rows
            if h1_values(row.get("metadata-h1-contents"))
            and h1_counts[" | ".join(h1_values(row.get("metadata-h1-contents"))).lower()]
            > 5
        ]
    elif issue == "Noindex present":
        matched_rows = [row for row in rows if noindex_enabled(row)]
    elif issue == "Nofollow present":
        matched_rows = [row for row in rows if nofollow_enabled(row)]
    elif issue == "Missing self-referencing canonical":
        matched_rows = [
            row for row in rows if not noindex_enabled(row) and not self_canonical_enabled(row)
        ]
    elif issue == "Sitemap/index-state mismatch":
        matched_rows = [row for row in rows if not sitemap_compliant(row)]
    elif issue == "No unique inlinks":
        matched_rows = [
            row for row in rows if (parse_int(row.get("No. of Unique Inlinks")) or 0) == 0
        ]
    elif issue == "Below inlink target":
        matched_rows = [row for row in rows if not inlink_target_met(row)]
    elif issue == "Crawl depth too deep":
        matched_rows = [row for row in rows if not depth_ok(row)]
    elif issue == "No internal outlinks":
        matched_rows = [
            row
            for row in rows
            if (parse_int(row.get("No. of Unique Outlinks to Internal Pages")) or 0) == 0
        ]
    elif issue == "Below outlink target":
        matched_rows = [row for row in rows if not outlink_target_met(row)]
    elif issue == "Speed measurement missing":
        matched_rows = [row for row in rows if speed_measured_health(row) is None]
    elif issue in {"LCP is not good", "CLS is not good", "INP is not good"}:
        metric = issue.split(" ", 1)[0]
        matched_rows = [
            row
            for row in rows
            if row_has_metric(row, metric) and row_metric_quality(row, metric) != "good"
        ]
    elif issue == "INP measurement missing":
        matched_rows = [row for row in rows if not row_has_metric(row, "INP")]
    elif issue == "Botify audience-keyword potential present":
        matched_rows = [row for row in rows if (botify_audience_potential(row) or 0) > 0]
    elif category == "Ranking Keywords":
        keyword_urls = keyword_urls_for_page_type(keyword_by_url, page_type)
        if issue == "Ranking URL is not in Botify crawl":
            matching_keywords = [
                (url, data) for url, data in keyword_urls.items() if url not in botify_urls
            ]
        elif issue == "No top-10 ranking keywords":
            matching_keywords = [
                (url, data) for url, data in keyword_urls.items() if data["top10"] == 0
            ]
        elif issue == "Has quick-win keywords in positions 11-20":
            matching_keywords = [
                (url, data)
                for url, data in keyword_urls.items()
                if keyword_quick_win_count(data) > 0
            ]
        elif issue == "Has striking-distance keywords in positions 21-50":
            matching_keywords = [
                (url, data)
                for url, data in keyword_urls.items()
                if keyword_striking_distance_count(data) > 0
            ]
        elif issue == "Has lost or unranked keywords":
            matching_keywords = [
                (url, data) for url, data in keyword_urls.items() if data["lost"] > 0
            ]
        else:
            matching_keywords = []
        return [
            keyword_detail_row(
                url,
                data,
                category,
                issue,
                recommendation,
                botify_by_url,
                main_keyword_by_url,
            )
            for url, data in matching_keywords
        ]
    else:
        matched_rows = []

    return [
        botify_detail_row(
            row,
            category,
            issue,
            recommendation,
            keyword_by_url,
            main_keyword_by_url,
        )
        for row in matched_rows
    ]


def write_issue_detail_files(
    detail_dir,
    issue_rows,
    botify_by_type,
    keyword_by_url,
    botify_urls,
    title_counts,
    meta_counts,
    h1_counts,
    main_keyword_by_url,
):
    detail_dir.mkdir(parents=True, exist_ok=True)
    for old_file in detail_dir.glob("*.csv"):
        old_file.unlink()

    for issue_row in issue_rows:
        detail_rows = issue_detail_rows(
            issue_row,
            botify_by_type,
            keyword_by_url,
            botify_urls,
            title_counts,
            meta_counts,
            h1_counts,
            main_keyword_by_url,
        )
        if not detail_rows:
            issue_row["detail_file"] = ""
            continue
        filename = issue_detail_filename(
            issue_row["page_type"], issue_row["category"], issue_row["issue"]
        )
        issue_row["detail_file"] = f"/outputs/issue_details/{filename}"
        with (detail_dir / filename).open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=ISSUE_DETAIL_FIELDS, lineterminator="\n"
            )
            writer.writeheader()
            writer.writerows(detail_rows)


def write_issue_summary(
    output_path,
    botify_by_type,
    keyword_by_url,
    botify_urls,
    title_counts,
    meta_counts,
    h1_counts,
    main_keyword_by_url,
):
    issue_rows = []

    for page_type in PAGE_TYPES:
        rows = [
            row for row in botify_by_type.get(page_type, []) if not is_bad_crawl(row)
        ]
        total = len(rows)
        targeted_rows = [
            row for row in rows if row["_normalized_url"] in main_keyword_by_url
        ]
        targeted_total = len(targeted_rows)

        add_issue(
            issue_rows,
            page_type,
            "Meta",
            "Missing page title",
            count_rows(rows, lambda row: not text(row.get("Title"))),
            total,
            "Botify Title field is empty.",
            "Add a unique, descriptive title for each eligible page.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Meta",
            "Page title length is not ideal",
            count_rows(
                rows,
                lambda row: text(row.get("Title"))
                and not (30 <= title_length(row) <= 65),
            ),
            total,
            "Ideal title length is 30-65 characters.",
            "Rewrite titles that are too short or too long while keeping the main topic clear.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Meta",
            "Duplicate page title",
            count_rows(
                rows,
                lambda row: text(row.get("Title"))
                and title_counts[text(row.get("Title")).lower()] > 1,
            ),
            total,
            "Same title appears on more than one crawled URL.",
            "Make titles unique at the page/template level.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Meta",
            "Missing meta description",
            count_rows(rows, lambda row: not text(row.get("Meta Description"))),
            total,
            "Botify Meta Description field is empty.",
            "Add a concise description for each eligible page.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Meta",
            "Meta description length is not ideal",
            count_rows(
                rows,
                lambda row: text(row.get("Meta Description"))
                and not (70 <= meta_description_length(row) <= 170),
            ),
            total,
            "Ideal meta description length is 70-170 characters.",
            "Rewrite descriptions that are too short or too long.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Meta",
            "Duplicate meta description",
            count_rows(
                rows,
                lambda row: text(row.get("Meta Description"))
                and meta_counts[text(row.get("Meta Description")).lower()] > 1,
            ),
            total,
            "Same meta description appears on more than one crawled URL.",
            "Make descriptions unique at the page/template level.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Meta",
            "Title does not align to main ranking keyword",
            count_rows(
                targeted_rows,
                lambda row: keyword_match_quality(
                    row.get("Title"), main_keyword_by_url.get(row["_normalized_url"])
                )
                not in {"full", "partial"},
            ),
            targeted_total,
            "Conductor main keyword is missing or weakly represented in the title.",
            "Update title templates so the primary ranking phrase is naturally represented.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Meta",
            "Meta description does not align to main ranking keyword",
            count_rows(
                targeted_rows,
                lambda row: keyword_match_quality(
                    row.get("Meta Description"),
                    main_keyword_by_url.get(row["_normalized_url"]),
                )
                not in {"full", "partial"},
            ),
            targeted_total,
            "Conductor main keyword is missing or weakly represented in the meta description.",
            "Update meta-description templates to reflect the primary ranking phrase.",
        )

        if page_type == "PDP":
            schema_checks = [
                ("Product schema missing", "Product Exists", "Add or fix Product schema."),
                ("Offer schema missing", "Offer Exists", "Add or fix Offer schema."),
                ("Breadcrumb schema missing", "Breadcrumb Exists", "Add breadcrumb schema."),
                ("Rating schema missing", "Rating Exists", "Add rating fields where eligible."),
                ("FAQ schema missing", "FAQ Exists", "Add FAQ schema where available; this is a nice-to-have."),
            ]
        elif page_type in {"Browse", "HVF"}:
            schema_checks = [
                ("Breadcrumb schema missing", "Breadcrumb Exists", "Add breadcrumb schema."),
                ("FAQ schema missing", "FAQ Exists", "Add FAQ schema where available; this is a nice-to-have."),
            ]
        elif page_type == "Editorial":
            schema_checks = []
            add_issue(
                issue_rows,
                page_type,
                "Schema",
                "Editorial schema implementation needs validation",
                total,
                total,
                "Editorial schema is fixed at 90 because article schema is not confirmed in the head.",
                "Validate Article schema placement and required properties before scoring Editorial at 100.",
                severity="medium",
            )
        else:
            schema_checks = [
                ("Breadcrumb schema missing", "Breadcrumb Exists", "Add breadcrumb schema."),
            ]
        for issue, column, recommendation in schema_checks:
            add_issue(
                issue_rows,
                page_type,
                "Schema",
                issue,
                count_rows(rows, lambda row, col=column: parse_bool(row.get(col)) is not True),
                total,
                f"Botify {column} is not true.",
                recommendation,
            )

        add_issue(
            issue_rows,
            page_type,
            "URL",
            "URL does not match the expected stable ID pattern",
            count_rows(rows, lambda row: stable_url_id_points(row) < 10),
            total,
            "Expected Chewy template identifier is missing or not recognized.",
            "Normalize URLs to the approved page-type pattern.",
        )
        add_issue(
            issue_rows,
            page_type,
            "URL",
            "URL slug is too generic",
            count_rows(
                rows,
                lambda row: len(
                    semantic_tokens(human_slug_for(row["_normalized_url"]).replace("-", " "))
                )
                < 2,
            ),
            total,
            "URL slug has fewer than two meaningful topic tokens.",
            "Use descriptive slugs that carry the page topic.",
        )
        add_issue(
            issue_rows,
            page_type,
            "URL",
            "URL slug does not align to main keyword",
            count_rows(
                targeted_rows,
                lambda row: url_intent_match_quality(row, main_keyword_by_url)
                not in {"full", "partial"},
            ),
            targeted_total,
            "Slug has weak or missing overlap with the Conductor main keyword.",
            "Update URL patterns so the slug reflects the primary search intent.",
        )
        add_issue(
            issue_rows,
            page_type,
            "URL",
            "URL path is too long or deep",
            count_rows(rows, path_too_long_or_deep),
            total,
            "Path length exceeds 115 characters or more than four path parts.",
            "Simplify URL hierarchy where the template allows it.",
        )

        add_issue(
            issue_rows,
            page_type,
            "H1",
            "Missing H1",
            count_rows(rows, lambda row: not h1_values(row.get("metadata-h1-contents"))),
            total,
            "Botify H1 field is empty.",
            "Add one descriptive H1 to the page template.",
        )
        add_issue(
            issue_rows,
            page_type,
            "H1",
            "Multiple H1s present",
            count_rows(rows, lambda row: len(h1_values(row.get("metadata-h1-contents"))) > 1),
            total,
            "Botify found more than one H1.",
            "Keep one primary H1 and move secondary headings to H2/H3.",
        )
        add_issue(
            issue_rows,
            page_type,
            "H1",
            "H1 is too generic",
            count_rows(
                rows,
                lambda row: h1_values(row.get("metadata-h1-contents"))
                and len(semantic_tokens(h1_text_for(row))) < 2,
            ),
            total,
            "H1 has fewer than two meaningful topic tokens.",
            "Make the H1 more descriptive of the page intent.",
        )
        add_issue(
            issue_rows,
            page_type,
            "H1",
            "H1 does not align to main keyword or topic",
            count_rows(
                rows,
                lambda row: h1_target_match_quality(row, main_keyword_by_url)
                not in {"full", "partial"},
            ),
            total,
            "H1 has weak or missing overlap with the target keyword or title topic.",
            "Update H1 templates to reflect the primary page intent.",
        )
        add_issue(
            issue_rows,
            page_type,
            "H1",
            "Duplicate H1 pattern",
            count_rows(
                rows,
                lambda row: h1_values(row.get("metadata-h1-contents"))
                and h1_counts[" | ".join(h1_values(row.get("metadata-h1-contents"))).lower()]
                > 5,
            ),
            total,
            "Same H1 appears across more than five crawled URLs.",
            "Make H1s more specific by page/template.",
        )

        add_issue(
            issue_rows,
            page_type,
            "Index",
            "Noindex present",
            count_rows(rows, noindex_enabled),
            total,
            "X-Robots-Tag noindex is true.",
            "Review intended index state and remove noindex from eligible indexable pages.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Index",
            "Nofollow present",
            count_rows(rows, nofollow_enabled),
            total,
            "X-Robots-Tag nofollow is true.",
            "Remove nofollow unless the page is intentionally excluded from link discovery.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Index",
            "Missing self-referencing canonical",
            count_rows(
                rows,
                lambda row: not noindex_enabled(row) and not self_canonical_enabled(row),
            ),
            total,
            "Canonical Points to Self is not true for an indexable URL.",
            "Set the canonical to the normalized self URL for eligible pages.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Index",
            "Sitemap/index-state mismatch",
            count_rows(rows, lambda row: not sitemap_compliant(row)),
            total,
            "Indexable URLs should be in sitemap; noindex URLs should be out of sitemap.",
            "Align sitemap inclusion to the intended index state.",
        )

        add_issue(
            issue_rows,
            page_type,
            "Links",
            "No unique inlinks",
            count_rows(rows, lambda row: (parse_int(row.get("No. of Unique Inlinks")) or 0) == 0),
            total,
            "Botify unique inlinks count is 0.",
            "Add discoverable internal links from relevant crawl paths.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Links",
            "Below inlink target",
            count_rows(rows, lambda row: not inlink_target_met(row)),
            total,
            "Unique inlink count is below the page-type target.",
            "Increase internal links from relevant navigation, PLPs, PDPs, or editorial surfaces.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Links",
            "Crawl depth too deep",
            count_rows(rows, lambda row: not depth_ok(row)),
            total,
            "Botify crawl depth exceeds the page-type threshold.",
            "Move important URLs closer to major crawl paths.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Links",
            "No internal outlinks",
            count_rows(
                rows,
                lambda row: (parse_int(row.get("No. of Unique Outlinks to Internal Pages")) or 0)
                == 0,
            ),
            total,
            "Botify unique internal outlinks count is 0.",
            "Add useful internal links to related pages, products, or supporting content.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Links",
            "Below outlink target",
            count_rows(rows, lambda row: not outlink_target_met(row)),
            total,
            "Unique internal outlink count is below the page-type target.",
            "Improve template linking to related destinations.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Links",
            "Breadcrumb/navigation signal missing",
            count_rows(rows, lambda row: parse_bool(row.get("Breadcrumb Exists")) is not True),
            total,
            "Breadcrumb Exists is not true.",
            "Add breadcrumb markup and visible navigation where appropriate.",
        )

        measured_speed_rows = [row for row in rows if speed_measured_health(row) is not None]
        add_issue(
            issue_rows,
            page_type,
            "Speed",
            "Speed measurement missing",
            total - len(measured_speed_rows),
            total,
            "No LCP, CLS, or INP metric is available for these URLs.",
            "Expand CrUX/PageSpeed/Botify field coverage; missing measurements are excluded from the score.",
            severity="medium",
        )
        for metric in ("LCP", "CLS", "INP"):
            metric_rows = [row for row in rows if row_has_metric(row, metric)]
            add_issue(
                issue_rows,
                page_type,
                "Speed",
                f"{metric} is not good",
                count_rows(metric_rows, lambda row, m=metric: row_metric_quality(row, m) != "good"),
                len(metric_rows),
                f"{metric} is measured but not in the good band.",
                "Prioritize template performance work for pages with needs-improvement or poor field data.",
            )
            if metric == "INP":
                add_issue(
                    issue_rows,
                    page_type,
                    "Speed",
                    "INP measurement missing",
                    total - len(metric_rows),
                    total,
                    "The current upload does not include INP for these URLs.",
                    "Add INP from CrUX, PageSpeed, or Botify exports.",
                    severity="medium",
                )

        keyword_urls = keyword_urls_for_page_type(keyword_by_url, page_type)
        keyword_total = len(keyword_urls)
        add_issue(
            issue_rows,
            page_type,
            "Ranking Keywords",
            "Ranking URL is not in Botify crawl",
            sum(1 for url in keyword_urls if url not in botify_urls),
            keyword_total,
            "Conductor has ranking URLs that do not match the Botify crawl.",
            "Review crawl coverage, URL normalization, and whether ranking URLs should be in the crawl denominator.",
            severity="medium",
        )
        add_issue(
            issue_rows,
            page_type,
            "Ranking Keywords",
            "No top-10 ranking keywords",
            sum(1 for data in keyword_urls.values() if data["top10"] == 0),
            keyword_total,
            "URL has tracked keywords but none rank in positions 1-10.",
            "Prioritize content, metadata, linking, and relevance improvements for this template cohort.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Ranking Keywords",
            "Has quick-win keywords in positions 11-20",
            sum(
                1
                for data in keyword_urls.values()
                if max(0, data["top20"] - data["top10"]) > 0
            ),
            keyword_total,
            "URL has at least one keyword ranking in positions 11-20.",
            "Push near-page-one terms with title/H1/meta/linking improvements.",
            severity="medium",
        )
        add_issue(
            issue_rows,
            page_type,
            "Ranking Keywords",
            "Has striking-distance keywords in positions 21-50",
            sum(
                1
                for data in keyword_urls.values()
                if max(0, data["top50"] - data["top20"]) > 0
            ),
            keyword_total,
            "URL has at least one keyword ranking in positions 21-50.",
            "Evaluate whether the page needs stronger content, intent coverage, or internal links.",
        )
        add_issue(
            issue_rows,
            page_type,
            "Ranking Keywords",
            "Has lost or unranked keywords",
            sum(1 for data in keyword_urls.values() if data["lost"] > 0),
            keyword_total,
            "URL has tracked keywords that currently do not rank.",
            "Review lost terms for relevance, indexability, and SERP intent fit.",
        )

        add_issue(
            issue_rows,
            page_type,
            "Ranking Potential",
            "Botify audience-keyword potential present",
            count_rows(rows, lambda row: (botify_audience_potential(row) or 0) > 0),
            total,
            "Botify reports keywords needed for the URL to achieve 90% audience.",
            "Use this cohort to prioritize pages with the largest reachable audience gap.",
            severity="medium",
        )

    issue_rows.sort(
        key=lambda row: (
            PAGE_TYPES.index(row["page_type"])
            if row["page_type"] in PAGE_TYPES
            else len(PAGE_TYPES),
            CATEGORY_ORDER.get(row["category"], 99),
            -int(row["affected_urls"]),
            row["issue"],
        )
    )
    write_issue_detail_files(
        output_path.parent / "issue_details",
        issue_rows,
        botify_by_type,
        keyword_by_url,
        botify_urls,
        title_counts,
        meta_counts,
        h1_counts,
        main_keyword_by_url,
    )
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=ISSUE_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(issue_rows)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    botify_rows, title_counts, meta_counts, h1_counts = read_botify()
    botify_urls = {
        row["_normalized_url"] for row in botify_rows if not is_bad_crawl(row)
    }
    botify_by_url = {row["_normalized_url"]: row for row in botify_rows}
    keyword_by_url, main_keyword_by_url = read_keywords()
    max_url_audience_potential = max(
        [botify_audience_potential(row) or 0 for row in botify_rows] or [0]
    )

    url_output = OUT_DIR / "url_keyword_summary.csv"
    with url_output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(
            [
                "normalized_url",
                "page_type",
                "matched_botify",
                "keyword_rows",
                "distinct_ranking_keywords",
                "current_ranked_keywords",
                "lost_or_unranked_keywords",
                "avg_current_rank",
                "top3_keywords",
                "top10_keywords",
                "top20_keywords",
                "top50_keywords",
                "top100_keywords",
                "ranking_keyword_score",
                "keywords_not_top10",
                "quick_win_11_20_keywords",
                "striking_distance_21_50_keywords",
                "botify_audience_keywords_to_90pct",
                "ranking_potential_score",
                "main_keyword",
                "title_matches_main_keyword",
                "meta_matches_main_keyword",
                "h1_matches_main_keyword",
                "url_matches_main_keyword",
                "keyword_groups_top",
            ]
        )
        for url, data in sorted(keyword_by_url.items()):
            rank_average = data["rank_sum"] / data["rank_n"] if data["rank_n"] else ""
            main_keyword = main_keyword_by_url.get(url, "")
            botify_row = botify_by_url.get(url)
            title_match = (
                keyword_match_quality(botify_row.get("Title"), main_keyword)
                if botify_row and main_keyword
                else ""
            )
            meta_match = (
                keyword_match_quality(botify_row.get("Meta Description"), main_keyword)
                if botify_row and main_keyword
                else ""
            )
            h1_match = (
                h1_target_match_quality(botify_row, main_keyword_by_url)
                if botify_row and main_keyword
                else ""
            )
            url_match = (
                url_intent_match_quality(botify_row, main_keyword_by_url)
                if botify_row and main_keyword
                else ""
            )
            botify_audience_keywords = (
                botify_audience_potential(botify_row) if botify_row else None
            )
            keyword_potential = keyword_opportunity_score(data)
            audience_potential_index = log_index(
                botify_audience_keywords, max_url_audience_potential
            )
            if keyword_potential is None and botify_audience_keywords is None:
                ranking_potential = None
            elif botify_audience_keywords is None:
                ranking_potential = keyword_potential
            else:
                ranking_potential = 0.75 * (keyword_potential or 0) + 0.25 * audience_potential_index
            writer.writerow(
                [
                    url,
                    page_type_for(url),
                    url in botify_urls,
                    data["rows"],
                    len(data["keywords"]),
                    data["current_ranked"],
                    data["lost"],
                    round(rank_average, 2) if rank_average != "" else "",
                    data["top3"],
                    data["top10"],
                    data["top20"],
                    data["top50"],
                    data["top100"],
                    round(keyword_score(data) or 0, 1),
                    data["rows"] - data["top10"],
                    max(0, data["top20"] - data["top10"]),
                    max(0, data["top50"] - data["top20"]),
                    (
                        ""
                        if botify_audience_keywords is None
                        else round(botify_audience_keywords, 1)
                    ),
                    (
                        ""
                        if ranking_potential is None
                        else round(ranking_potential, 1)
                    ),
                    main_keyword,
                    title_match,
                    meta_match,
                    h1_match,
                    url_match,
                    "; ".join(group for group, _ in data["groups"].most_common(5)),
                ]
            )

    botify_by_type = defaultdict(list)
    for row in botify_rows:
        botify_by_type[row["_page_type"]].append(row)

    keywords_by_type = defaultdict(
        lambda: {
            "urls": set(),
            "matched_urls": set(),
            "rows": 0,
            "keywords": set(),
            "current_ranked": 0,
            "lost": 0,
            "top3": 0,
            "top10": 0,
            "top20": 0,
            "top50": 0,
            "top100": 0,
            "score_sum": 0.0,
            "rank_sum": 0.0,
            "rank_n": 0,
        }
    )
    for url, data in keyword_by_url.items():
        page_type = page_type_for(url)
        if page_type not in PAGE_TYPES:
            continue
        summary = keywords_by_type[page_type]
        summary["urls"].add(url)
        if url in botify_urls:
            summary["matched_urls"].add(url)
        summary["rows"] += data["rows"]
        summary["keywords"].update(data["keywords"])
        for key in [
            "current_ranked",
            "lost",
            "top3",
            "top10",
            "top20",
            "top50",
            "top100",
        ]:
            summary[key] += data[key]
        summary["score_sum"] += data["score_sum"]
        summary["rank_sum"] += data["rank_sum"]
        summary["rank_n"] += data["rank_n"]

    audience_potential_by_type = {
        page_type: numeric_distribution(
            [
                botify_audience_potential(row)
                for row in botify_by_type.get(page_type, [])
                if not is_bad_crawl(row)
            ]
        )
        for page_type in PAGE_TYPES
    }
    maxes = {
        "quick_win": max(
            [
                (
                    max(0, keywords_by_type[page_type]["top20"] - keywords_by_type[page_type]["top10"])
                    / keywords_by_type[page_type]["rows"]
                    * 100
                )
                if keywords_by_type[page_type]["rows"]
                else 0
                for page_type in PAGE_TYPES
            ]
            or [0]
        ),
        "striking_distance": max(
            [
                (
                    max(0, keywords_by_type[page_type]["top50"] - keywords_by_type[page_type]["top20"])
                    / keywords_by_type[page_type]["rows"]
                    * 100
                )
                if keywords_by_type[page_type]["rows"]
                else 0
                for page_type in PAGE_TYPES
            ]
            or [0]
        ),
        "avg_audience": max(
            [
                audience_potential_by_type[page_type]["avg"] or 0
                for page_type in PAGE_TYPES
            ]
            or [0]
        ),
        "total_audience": max(
            [
                audience_potential_by_type[page_type]["total"]
                for page_type in PAGE_TYPES
            ]
            or [0]
        ),
    }

    page_output = OUT_DIR / "page_type_heatmap_with_ranking_keywords.csv"
    with page_output.open("w", newline="", encoding="utf-8") as handle:
        fields = [
            "page_type",
            "botify_urls",
            "scorable_urls",
            "excluded_bad_crawl_urls",
            "excluded_bad_crawl_pct",
            "keyword_urls",
            "keyword_urls_matched_botify",
            "keyword_inventory_coverage_pct",
            "keyword_rows",
            "distinct_keywords",
            "avg_current_rank",
            "top3_share_pct",
            "top10_share_pct",
            "top20_share_pct",
            "lost_or_unranked_share_pct",
            "ranking_keywords",
            "keywords_not_top10_share_pct",
            "quick_win_11_20_share_pct",
            "striking_distance_21_50_share_pct",
            "botify_audience_potential_urls",
            "audience_potential_coverage_pct",
            "avg_botify_audience_potential",
            "total_botify_audience_potential",
            "ranking_potential",
            "main_keyword_target_coverage_pct",
            "h1_present_pct",
            "single_h1_pct",
            "h1_main_keyword_match_pct",
            "title_main_keyword_match_pct",
            "meta_main_keyword_match_pct",
            "url_main_keyword_match_pct",
            "index_controls_clear_pct",
            "noindex_pct",
            "self_canonical_pct",
            "sitemap_compliance_pct",
            "canonical_sitemap_compliance_pct",
            "median_inlinks",
            "inlink_target_met_pct",
            "median_outlinks",
            "outlink_target_met_pct",
            "crawl_depth_ok_pct",
            "link_breadcrumb_pct",
            "lcp_measured_pct",
            "lcp_good_pct",
            "cls_measured_pct",
            "cls_good_pct",
            "inp_measured_pct",
            "inp_good_pct",
            "speed_measured_health",
            "speed_scored_urls",
            "speed_coverage_pct",
            "meta",
            "schema",
            "url",
            "h1",
            "index_readiness",
            "links",
            "speed",
            "overall_proxy_excl_ranking_keywords",
            "overall_proxy_incl_ranking_keywords",
        ]
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for page_type in PAGE_TYPES:
            rows = botify_by_type.get(page_type, [])
            scorable_rows = [row for row in rows if not is_bad_crawl(row)]
            excluded_count = len(rows) - len(scorable_rows)
            keyword_summary = keywords_by_type[page_type]
            speeds = [speed_measured_health(row) for row in scorable_rows]
            speeds = [value for value in speeds if value is not None]
            speed_coverage = (
                len(speeds) / len(scorable_rows) * 100 if scorable_rows else 0
            )
            measured_speed = sum(speeds) / len(speeds) if speeds else None
            final_speed = measured_speed
            ranking_rows = keyword_summary["rows"]
            ranking_score = None
            if ranking_rows:
                ranking_score = min(
                    100,
                    0.75 * (keyword_summary["score_sum"] / ranking_rows)
                    + 0.15 * (keyword_summary["top10"] / ranking_rows * 100)
                    + 0.10
                    * (keyword_summary["current_ranked"] / ranking_rows * 100),
                )
            audience_summary = audience_potential_by_type[page_type]
            potential_score = ranking_potential_score(
                keyword_summary, audience_summary, maxes
            )
            targeted_rows = [
                row
                for row in scorable_rows
                if row["_normalized_url"] in main_keyword_by_url
            ]
            title_keyword_matches = [
                keyword_match_quality(
                    row.get("Title"), main_keyword_by_url.get(row["_normalized_url"])
                )
                for row in targeted_rows
            ]
            meta_keyword_matches = [
                keyword_match_quality(
                    row.get("Meta Description"),
                    main_keyword_by_url.get(row["_normalized_url"]),
                )
                for row in targeted_rows
            ]
            h1_keyword_matches = [
                h1_target_match_quality(row, main_keyword_by_url)
                for row in targeted_rows
            ]
            url_keyword_matches = [
                url_intent_match_quality(row, main_keyword_by_url)
                for row in targeted_rows
            ]
            scores = {
                "meta": average(
                    scorable_rows,
                    lambda row: meta_score(
                        row, title_counts, meta_counts, main_keyword_by_url
                    ),
                ),
                "schema": average(scorable_rows, schema_score),
                "url": average(
                    scorable_rows, lambda row: url_score(row, main_keyword_by_url)
                ),
                "h1": average(
                    scorable_rows,
                    lambda row: h1_score(row, h1_counts, main_keyword_by_url),
                ),
                "index_readiness": average(scorable_rows, index_score),
                "links": average(scorable_rows, links_score),
                "speed": final_speed,
            }
            overall_parts = [value for value in scores.values() if value is not None]
            overall = sum(overall_parts) / len(overall_parts) if overall_parts else None
            with_ranking = overall_parts + ([ranking_score] if ranking_score is not None else [])
            average_rank = (
                keyword_summary["rank_sum"] / keyword_summary["rank_n"]
                if keyword_summary["rank_n"]
                else None
            )
            botify_count = len(rows)
            keyword_url_count = len(keyword_summary["urls"])
            writer.writerow(
                {
                    "page_type": page_type,
                    "botify_urls": botify_count,
                    "scorable_urls": len(scorable_rows),
                    "excluded_bad_crawl_urls": excluded_count,
                    "excluded_bad_crawl_pct": (
                        round(excluded_count / botify_count * 100, 1)
                        if botify_count
                        else 0
                    ),
                    "keyword_urls": keyword_url_count,
                    "keyword_urls_matched_botify": len(keyword_summary["matched_urls"]),
                    "keyword_inventory_coverage_pct": (
                        ""
                        if not botify_count
                        else round(keyword_url_count / botify_count * 100, 2)
                    ),
                    "keyword_rows": ranking_rows,
                    "distinct_keywords": len(keyword_summary["keywords"]),
                    "avg_current_rank": "" if average_rank is None else round(average_rank, 1),
                    "top3_share_pct": (
                        "" if not ranking_rows else round(keyword_summary["top3"] / ranking_rows * 100, 1)
                    ),
                    "top10_share_pct": (
                        "" if not ranking_rows else round(keyword_summary["top10"] / ranking_rows * 100, 1)
                    ),
                    "top20_share_pct": (
                        "" if not ranking_rows else round(keyword_summary["top20"] / ranking_rows * 100, 1)
                    ),
                    "lost_or_unranked_share_pct": (
                        "" if not ranking_rows else round(keyword_summary["lost"] / ranking_rows * 100, 1)
                    ),
                    "ranking_keywords": "" if ranking_score is None else round(ranking_score, 1),
                    "keywords_not_top10_share_pct": (
                        ""
                        if not ranking_rows
                        else round(
                            (ranking_rows - keyword_summary["top10"])
                            / ranking_rows
                            * 100,
                            1,
                        )
                    ),
                    "quick_win_11_20_share_pct": (
                        ""
                        if not ranking_rows
                        else round(
                            max(
                                0,
                                keyword_summary["top20"]
                                - keyword_summary["top10"],
                            )
                            / ranking_rows
                            * 100,
                            1,
                        )
                    ),
                    "striking_distance_21_50_share_pct": (
                        ""
                        if not ranking_rows
                        else round(
                            max(
                                0,
                                keyword_summary["top50"]
                                - keyword_summary["top20"],
                            )
                            / ranking_rows
                            * 100,
                            1,
                        )
                    ),
                    "botify_audience_potential_urls": audience_summary["count"],
                    "audience_potential_coverage_pct": (
                        round(
                            audience_summary["count"] / len(scorable_rows) * 100,
                            1,
                        )
                        if scorable_rows
                        else 0
                    ),
                    "avg_botify_audience_potential": (
                        ""
                        if audience_summary["avg"] is None
                        else round(audience_summary["avg"], 1)
                    ),
                    "total_botify_audience_potential": round(
                        audience_summary["total"], 1
                    ),
                    "ranking_potential": (
                        "" if potential_score is None else round(potential_score, 1)
                    ),
                    "main_keyword_target_coverage_pct": (
                        round(len(targeted_rows) / len(scorable_rows) * 100, 1)
                        if scorable_rows
                        else 0
                    ),
                    "h1_present_pct": (
                        round(
                            sum(
                                1
                                for row in scorable_rows
                                if h1_values(row.get("metadata-h1-contents"))
                            )
                            / len(scorable_rows)
                            * 100,
                            1,
                        )
                        if scorable_rows
                        else 0
                    ),
                    "single_h1_pct": (
                        round(
                            sum(
                                1
                                for row in scorable_rows
                                if len(h1_values(row.get("metadata-h1-contents")))
                                == 1
                            )
                            / len(scorable_rows)
                            * 100,
                            1,
                        )
                        if scorable_rows
                        else 0
                    ),
                    "h1_main_keyword_match_pct": (
                        ""
                        if not targeted_rows
                        else round(
                            sum(
                                1
                                for quality in h1_keyword_matches
                                if quality in {"full", "partial"}
                            )
                            / len(targeted_rows)
                            * 100,
                            1,
                        )
                    ),
                    "title_main_keyword_match_pct": (
                        ""
                        if not targeted_rows
                        else round(
                            sum(
                                1
                                for quality in title_keyword_matches
                                if quality in {"full", "partial"}
                            )
                            / len(targeted_rows)
                            * 100,
                            1,
                        )
                    ),
                    "meta_main_keyword_match_pct": (
                        ""
                        if not targeted_rows
                        else round(
                            sum(
                                1
                                for quality in meta_keyword_matches
                                if quality in {"full", "partial"}
                            )
                            / len(targeted_rows)
                            * 100,
                            1,
                        )
                    ),
                    "url_main_keyword_match_pct": (
                        ""
                        if not targeted_rows
                        else round(
                            sum(
                                1
                                for quality in url_keyword_matches
                                if quality in {"full", "partial"}
                            )
                            / len(targeted_rows)
                            * 100,
                            1,
                        )
                    ),
                    "index_controls_clear_pct": (
                        round(
                            sum(
                                1
                                for row in scorable_rows
                                if not noindex_enabled(row)
                                and not nofollow_enabled(row)
                            )
                            / len(scorable_rows)
                            * 100,
                            1,
                        )
                        if scorable_rows
                        else 0
                    ),
                    "noindex_pct": (
                        round(
                            sum(1 for row in scorable_rows if noindex_enabled(row))
                            / len(scorable_rows)
                            * 100,
                            1,
                        )
                        if scorable_rows
                        else 0
                    ),
                    "self_canonical_pct": (
                        round(
                            sum(
                                1
                                for row in scorable_rows
                                if self_canonical_enabled(row)
                            )
                            / len(scorable_rows)
                            * 100,
                            1,
                        )
                        if scorable_rows
                        else 0
                    ),
                    "sitemap_compliance_pct": (
                        round(
                            sum(1 for row in scorable_rows if sitemap_compliant(row))
                            / len(scorable_rows)
                            * 100,
                            1,
                        )
                        if scorable_rows
                        else 0
                    ),
                    "canonical_sitemap_compliance_pct": (
                        pct_rows(scorable_rows, canonical_sitemap_compliant)
                    ),
                    "median_inlinks": (
                        ""
                        if median_numeric(
                            scorable_rows, "No. of Unique Inlinks"
                        )
                        is None
                        else round(
                            median_numeric(
                                scorable_rows, "No. of Unique Inlinks"
                            ),
                            1,
                        )
                    ),
                    "inlink_target_met_pct": pct_rows(
                        scorable_rows, inlink_target_met
                    ),
                    "median_outlinks": (
                        ""
                        if median_numeric(
                            scorable_rows,
                            "No. of Unique Outlinks to Internal Pages",
                        )
                        is None
                        else round(
                            median_numeric(
                                scorable_rows,
                                "No. of Unique Outlinks to Internal Pages",
                            ),
                            1,
                        )
                    ),
                    "outlink_target_met_pct": pct_rows(
                        scorable_rows, outlink_target_met
                    ),
                    "crawl_depth_ok_pct": pct_rows(scorable_rows, depth_ok),
                    "link_breadcrumb_pct": pct_rows(
                        scorable_rows,
                        lambda row: parse_bool(row.get("Breadcrumb Exists")) is True,
                    ),
                    "lcp_measured_pct": metric_measured_pct(scorable_rows, "LCP"),
                    "lcp_good_pct": metric_good_pct(scorable_rows, "LCP"),
                    "cls_measured_pct": metric_measured_pct(scorable_rows, "CLS"),
                    "cls_good_pct": metric_good_pct(scorable_rows, "CLS"),
                    "inp_measured_pct": metric_measured_pct(scorable_rows, "INP"),
                    "inp_good_pct": metric_good_pct(scorable_rows, "INP"),
                    "speed_measured_health": (
                        "" if measured_speed is None else round(measured_speed, 1)
                    ),
                    "speed_scored_urls": len(speeds),
                    "speed_coverage_pct": (
                        round(speed_coverage, 1) if scorable_rows else 0
                    ),
                    **{
                        key: "" if value is None else round(value, 1)
                        for key, value in scores.items()
                    },
                    "overall_proxy_excl_ranking_keywords": (
                        "" if overall is None else round(overall, 1)
                    ),
                    "overall_proxy_incl_ranking_keywords": (
                        "" if not with_ranking else round(sum(with_ranking) / len(with_ranking), 1)
                    ),
                }
            )

    issue_output = OUT_DIR / "page_type_issue_summary.csv"
    write_issue_summary(
        issue_output,
        botify_by_type,
        keyword_by_url,
        botify_urls,
        title_counts,
        meta_counts,
        h1_counts,
        main_keyword_by_url,
    )

    print(page_output)
    print(url_output)
    print(issue_output)


if __name__ == "__main__":
    main()
